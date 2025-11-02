#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
予約情報をExcelに追記するスクリプト（テンプレート使用版・罫線修正）
"""

import sys
import json
import shutil
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

def get_project_paths():
    """プロジェクトのパスを取得"""
    project_root = Path(__file__).parent.parent
    
    template_dir = project_root / "data" / "excel" / "templates"
    base_dir = project_root / "data" / "excel"
    output_dir = project_root / "public" / "downloads"
    
    template_dir.mkdir(parents=True, exist_ok=True)
    base_dir.mkdir(parents=True, exist_ok=True)
    output_dir.mkdir(parents=True, exist_ok=True)
    
    return template_dir, base_dir, output_dir

def format_date(date_str):
    """日付を 'MM月DD日' 形式に変換"""
    try:
        date_obj = datetime.strptime(date_str, '%Y-%m-%d')
        return f"{date_obj.month}月{date_obj.day}日"
    except:
        return date_str

def update_excel(booking_data):
    """Excelファイルを更新（テンプレート使用）"""
    try:
        template_dir, base_dir, output_dir = get_project_paths()
        
        date_obj = datetime.strptime(booking_data['date'], '%Y-%m-%d')
        year = date_obj.year
        month = date_obj.month
        
        file_name = f"{year}年 {month:02d}月売上.xlsx"
        file_path = base_dir / file_name
        
        template_name = "月次売上テンプレート.xlsx"
        template_path = template_dir / template_name
        
        if not file_path.exists():
            if not template_path.exists():
                error_result = {
                    "success": False,
                    "error": f"テンプレートファイルが見つかりません: {template_path}"
                }
                print(json.dumps(error_result, ensure_ascii=False))
                return 1
            
            shutil.copy(template_path, file_path)
        
        wb = load_workbook(file_path)
        ws = wb.active
        
        # 新しい予約の日付
        new_date = datetime.strptime(booking_data['date'], '%Y-%m-%d')
        
        # 適切な挿入位置を探す（14行目から）
        insert_row = 14
        current_row = 14
        
        # 既存のデータを走査して挿入位置を決定
        while ws[f'A{current_row}'].value is not None:
            cell_value = ws[f'A{current_row}'].value
            
            # セルの値から日付を抽出（例: "10月27日" → datetime）
            try:
                if isinstance(cell_value, str) and '月' in cell_value and '日' in cell_value:
                    # "10月27日" 形式から日付を抽出
                    month_str = cell_value.split('月')[0]
                    day_str = cell_value.split('月')[1].replace('日', '')
                    existing_date = datetime(new_date.year, int(month_str), int(day_str))
                    
                    # 新しい日付が既存の日付より前なら、この位置に挿入
                    if new_date < existing_date:
                        insert_row = current_row
                        break
            except:
                pass
            
            current_row += 1
        
        # 挿入位置が既存データの中なら、行を挿入
        if ws[f'A{insert_row}'].value is not None:
            ws.insert_rows(insert_row)
        
        # データを書き込み
        ws[f'A{insert_row}'] = format_date(booking_data['date'])
        ws[f'C{insert_row}'] = booking_data['customer_name']
        ws[f'E{insert_row}'] = booking_data['staff_name']
        ws[f'H{insert_row}'] = str(booking_data['visit_count']) + "回目"
        
        # ★修正: 罫線を細線(thin)に統一
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # スタイル適用
        for col in ['A', 'C', 'E', 'H']:
            source_cell = ws[f'{col}13']
            target_cell = ws[f'{col}{insert_row}']
            
            if source_cell.font:
                target_cell.font = Font(
                    name=source_cell.font.name,
                    size=source_cell.font.size,
                    bold=source_cell.font.bold,
                    color=source_cell.font.color
                )
            
            if source_cell.alignment:
                target_cell.alignment = Alignment(
                    horizontal=source_cell.alignment.horizontal,
                    vertical=source_cell.alignment.vertical
                )
            
            # ★修正: 罫線は細線を強制適用
            target_cell.border = thin_border
        
        # 保存
        wb.save(file_path)
        
        # ダウンロード用にコピー
        output_path = output_dir / file_name
        wb.save(output_path)
        
        result = {
            "success": True,
            "file_name": file_name,
            "row": insert_row,
            "file_path": str(file_path),
            "output_path": str(output_path)
        }
        
        print(json.dumps(result, ensure_ascii=False))
        return 0
        
    except Exception as e:
        error_result = {
            "success": False,
            "error": str(e)
        }
        print(json.dumps(error_result, ensure_ascii=False))
        return 1

def main():
    if len(sys.argv) < 2:
        error_result = {
            "success": False,
            "error": "引数が不足しています"
        }
        print(json.dumps(error_result, ensure_ascii=False))
        return 1
    
    try:
        booking_data = json.loads(sys.argv[1])
        return update_excel(booking_data)
    except json.JSONDecodeError as e:
        error_result = {
            "success": False,
            "error": f"JSON解析エラー: {str(e)}"
        }
        print(json.dumps(error_result, ensure_ascii=False))
        return 1

if __name__ == "__main__":
    sys.exit(main())